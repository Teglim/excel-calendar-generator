import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from datetime import date
import jpholiday
import unicodedata

import os
import time
import win32com.client
from PIL import ImageGrab

def width(s):
    w = 0
    for c in s:
        if unicodedata.east_asian_width(c) in "FWA":
            w += 2
        else:
            w += 1
    return w

thin = Side(style="thin")

def set_border(ws, r0, c):
    for i in range(4):
        cell = ws.cell(row=r0+i, column=c)

        cell.border = Border(
            top=thin if i == 0 else None,
            bottom=thin if i == 3 else None,
            left=thin,
            right=thin
        )

def px_to_pt(px):
    return px * 0.75

def format_text(name, sub, W=15):
    if not sub:
        return name

    # 長さ制限（表示幅ベース）
    def trim(s, W):
        res = ""
        for c in s:
            if width(res + c) > W:
                break
            res += c
        return res

    name = trim(name, W)
    sub = trim(sub, W)

    w_name = width(name)
    w_sub = width(sub)

    space = W - w_name - w_sub

    if space < 1:
        return name

    return name + " " * space + sub

def save_excel_as_image(xlsx_path, out_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
    ws = wb.Sheets(1)

    used = ws.UsedRange

    # 範囲全体を画像コピー
    used.CopyPicture(
        Appearance=1,   # xlScreen
        Format=2        # xlBitmap
    )

    for _ in range(20):
        time.sleep(0.1)
        img = ImageGrab.grabclipboard()
        if img is not None:
            break
    
    if img is None:
        wb.Close(False)
        excel.Quit()
        raise RuntimeError("画像取得失敗")

    img.save(out_path)

    wb.Close(False)
    excel.Quit()

def cal():
    # --- 入力 ---
    def read_input():
        Y = M = 0
        ev = {}
        opt = ""
        mode = ""
        with open("input.txt", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                if line.startswith("<<"):
                    mode = line[3:]
                    continue
                if mode == "Y":
                    Y = int(line)
                elif mode == "M":
                    M = int(line)
                elif mode == "D":
                    sp = [x.strip() for x in line.split(":")]
                    d = int(sp[0])
                    name = sp[1]
                    sub = sp[2] if len(sp) > 2 else ""
                    if d not in ev:
                        ev[d] = []
                    ev[d].append((name, sub))

                # 未実装
                # elif mode == "OPT":
                #     opt += line + "\n"
        return Y, M, ev

    def read_setting():
        mp = {}
        mode = ""
        with open("setting.txt", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                if line.startswith("<<"):
                    mode = line[3:]
                    continue
                if mode == "COLOR":
                    name, col = [x.strip() for x in line.split(":")]
                    mp[name] = col.replace("#", "")
        return mp

    Y, M, ev = read_input()
    color = read_setting()

    wb = Workbook()
    ws = wb.active

    # --- フォント ---
    font = Font(name="ＭＳ ゴシック", size=18)

    # --- 列幅 ---
    ws.column_dimensions["A"].width = 50 / 7
    for c in range(2, 9):
        ws.column_dimensions[chr(ord('A')+c-1)].width = 170 / 7

    # --- 行高 ---
    heights = [20, 60, 10, 35]
    for i, h in enumerate(heights, 1):
        ws.row_dimensions[i].height = px_to_pt(h)

    # --- タイトル ---
    # D：年度（右）
    cell = ws.cell(row=2, column=4, value=Y)
    cell.font = Font(name="ＭＳ ゴシック", size=22)
    cell.alignment = Alignment(horizontal="right", vertical="center")

    # E：月（中央・大きめ）
    cell = ws.cell(row=2, column=5, value=M)
    cell.font = Font(name="ＭＳ ゴシック", size=36)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # F：英語（左）
    cell = ws.cell(row=2, column=6, value=calendar.month_name[M])
    cell.font = Font(name="ＭＳ ゴシック", size=22)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- 曜日 ---
    week = ["Sun","Mon","Tue","Wed","Thu","Fri","Sat"]

    for i in range(7):
        c = i + 2
        cell = ws.cell(row=4, column=c, value=week[i])

        # フォント（白）
        cell.font = Font(name="ＭＳ ゴシック", size=14, color="FFFFFF")

        # 配置
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 背景色
        if i == 0:
            col = "FF0000"  # 日曜
        elif i == 6:
            col = "0000FF"  # 土曜
        else:
            col = "333333"  # 平日

        cell.fill = PatternFill(
            start_color=col,
            end_color=col,
            fill_type="solid"
        )

    # --- カレンダー ---
    calendar.setfirstweekday(calendar.SUNDAY)
    cal = calendar.monthcalendar(Y, M)

    base_row = 5

    for w in cal:
        for i in range(7):
            d = w[i]
            if d == 0:
                continue

            col = i + 2
            r0 = base_row
            for k in range(4):
                ws.row_dimensions[r0 + k].height = px_to_pt(30)

            # --- 色判定 ---
            is_sun = (i == 0)
            is_sat = (i == 6)
            is_hol = jpholiday.is_holiday(date(Y, M, d))

            bg = None
            if is_sun or is_hol:
                bg = "FFD9D9"   # 薄い赤
            elif is_sat:
                bg = "D9E1FF"   # 薄い青

            # --- a行（日付） ---
            cell = ws.cell(row=r0, column=col, value=d)
            cell.font = Font(name="ＭＳ ゴシック", size=18)
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # --- b,c,d ---
            for k in range(1, 4):
                ws.cell(row=r0+k, column=col).font = Font(name="ＭＳ ゴシック", size=18)

            # --- 背景色（4行全部） ---
            if bg:
                for k in range(4):
                    cell = ws.cell(row=r0+k, column=col)
                    cell.fill = PatternFill(
                        start_color=bg,
                        end_color=bg,
                        fill_type="solid"
                    )

            # --- 予定 ---
            if d in ev:
                if len(ev[d]) > 3:
                    print(f"Error: {d}日に予定が多すぎます")

                for idx, (name, sub) in enumerate(ev[d][:3]):
                    txt = name
                    if sub:
                        txt = format_text(name, sub)

                    cell = ws.cell(row=r0+1+idx, column=col, value=txt)
                    cell.font = Font(name="ＭＳ ゴシック", size=18)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                    # イベント色は上書き
                    if name in color:
                        cell.fill = PatternFill(
                            start_color=color[name],
                            end_color=color[name],
                            fill_type="solid"
                        )

            # --- 枠線 ---
            set_border(ws, r0, col)

        base_row += 4

    ws.sheet_view.showGridLines = False
    
    fname = f"{Y}_{M:02d}_cal.xlsx"
    wb.save(fname)

    img_name = f"{Y}_{M:02d}_cal.png"
    save_excel_as_image(fname, img_name)

    print("saved:", fname)
    print("saved:", img_name)

if __name__ == "__main__":
    cal()
